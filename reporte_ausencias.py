def procesa_reporte_ausentismo():
    import pandas as pd

    import openpyxl
    from openpyxl import Workbook
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.utils import column_index_from_string
    from openpyxl.styles import PatternFill
    from openpyxl.styles import Border, Side
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils.dataframe import dataframe_to_rows

    import plotly.express as px
    import matplotlib as plt

    import datetime as dt
    import time
    import re


    def selecciona_celdas(rango=None,
                          sheet=None,
                          value=None,
                          font=None,
                          alignment=None,
                          border=None,
                          fill=None,
                          number_format=None,
                          hyperlink=None,
                          comment=None,
                          rule= None):

        def letra_a_numero(columna):
            resultado = 0
            for i, letra in enumerate(reversed(columna)):
                resultado += (ord(letra.upper()) - ord('A') + 1) * (26 ** i)
            return resultado

        r = rango.split(":")
        row = []
        columns = []

        for i in range(0, 2):
            col_num = []
            row_num = []
            for caracter in r[i]:
                if caracter.isalpha():
                    letras = caracter
                    num = letra_a_numero(letras)
                    col_num.append(str(num))

                elif caracter.isdigit():
                    row_num.append(str(caracter))

            columns.append(''.join(col_num))
            row.append(''.join(row_num))

        col_start = int(columns[0])
        col_end = int(columns[1]) + 1
        row_start = int(row[0])
        row_end = int(row[1]) + 1

        for col in range(col_start, col_end):
            for row in range(row_start, row_end):
                cell = sheet.cell(row=row, column=col)
                if value is not None:
                    cell.value = value

                if font is not None:
                    cell.font = font

                if alignment is not None:
                    cell.alignment = alignment

                if border is not None:
                    cell.border = border

                if fill is not None:
                    cell.fill = fill

                if number_format is not None:
                    cell.number_format = number_format

                if hyperlink is not None:
                    cell.hyperlink = hyperlink

                if comment is not None:
                    cell.comment = comment

                if rule is not None:
                    cell.rule = rule


    def conv_num_a_letra(numero):
        letras = ""
        while numero > 0:
            numero -= 1
            letras = chr((numero % 26) + 65) + letras
            numero //= 26
        return letras

    directorio = r"C:\Users\bchavat\Desktop\Automat Reportes\ConstruccionAusencias"
    reporte_ausencias = pd.read_excel( directorio + "\Reporte Totales Horas con Autorizaciones.xls")
    reporte_busqueda_rapida = pd.read_excel( directorio + "\Búsqueda rápida.xlsx")
    cc_info = pd.read_excel(r"C:\Users\bchavat\Desktop\Automat Reportes\r_data\centros_de_costo_info.xlsx")
    codigos_de_pago = pd.read_excel(r"C:\Users\bchavat\Desktop\Automat Reportes\r_data\codigos_de_pago.xlsx")

    def procesa_arch_busqueda_rapida(archivo):
        col_sep = archivo["Unnamed: 5"].str.split('/', expand=True)
        archivo = pd.concat([archivo, col_sep], axis=1)

        archivo.drop(["Unnamed: 2", "Unnamed: 5", 1, 2, 4, 3, 6], axis=1, inplace=True)

        archivo = archivo.set_axis(['Nombre',
                                    'NroColab',
                                    'Centro de Costo',
                                    'Puesto',
                                    'Empresa',
                                    'Nochero'],
                                   axis=1
                                   )

        archivo = archivo.iloc[1:, :]

        archivo.reset_index(drop=True, inplace=True)
        archivo['NroColab'] = pd.to_numeric(archivo['NroColab'], downcast='integer',
                                            errors='coerce')

        print("Busqueda rapida -*OK*-")
        return archivo

    reporte_busqueda_rapida = procesa_arch_busqueda_rapida(reporte_busqueda_rapida)

    name_col = []
    max_col = int(reporte_ausencias.shape[1])
    for i in range(0, max_col):
        celda = reporte_ausencias.iloc[2, i]
        name_col.append(celda)

    reporte_ausencias = reporte_ausencias.iloc[3:, :]
    reporte_ausencias = reporte_ausencias.set_axis(name_col, axis=1)
    reporte_ausencias.reset_index(drop=True, inplace=True)
    reporte_ausencias["N° FUNC."] = reporte_ausencias["N° FUNC."].astype(int)


    reporte_ausencias = pd.merge(left=reporte_ausencias, right=codigos_de_pago,
                                 how="left", on="CODIGO DE PAGO")
    reporte_ausencias = reporte_ausencias.dropna(subset=["TIPO"])
    reporte_ausencias = pd.merge(left=reporte_ausencias, right=reporte_busqueda_rapida,
                                 how="left", left_on="N° FUNC.", right_on="NroColab")
    reporte_ausencias = pd.merge(left=reporte_ausencias, right=cc_info, how="left",
                                 left_on="Centro de Costo", right_on="SUCURSAL")

    reporte_ausencias = reporte_ausencias[["N° FUNC.",
                                           "NOMBRE",
                                           "Centro de Costo",
                                           "ZONA",
                                           "DEPARTAMENTO",
                                           "BARRIO/LOCALIDAD",
                                           "Gerente",
                                           "Puesto",
                                           "FECHA",
                                           "CODIGO DE PAGO",
                                           "HORAS",
                                           "TIPO",
                                           "SUPERVISOR"
                                           ]]
    reporte_ausencias = reporte_ausencias.rename(columns=
                                                 {'Centro de Costo': 'TIENDA',
                                                  'Gerente': 'Jefe Supervisor',
                                                  'Puesto': 'CARGO'})
    reporte_ausencias['FECHA'] = pd.to_datetime(reporte_ausencias['FECHA']).dt.strftime("%d-%m-%Y")

    reporte_ausencias.to_excel("base_ausencias.xlsx")


    lista_cc = ["FARMASHOP.*", "BELA.*", "SEO", "Ecommerce 900.*", "LOG.*"]
    cc = ["Farmashop", "Bela", "SEO", "Ecommerce 900", "Logistica"]

    workbook = Workbook()

    for lista_re, nombre in zip(lista_cc, cc):
        df_anexo = reporte_ausencias[reporte_ausencias["TIENDA"].str.contains(lista_re,
                                                                              regex=True)]
        df_anexo_for_tabla = df_anexo.copy()
        df_anexo_for_tabla.loc[:, "FECHA"] = pd.to_datetime(
                                df_anexo_for_tabla.loc[:, "FECHA"],format="%d-%m-%Y").dt.date
        df_anexo_for_tabla.sort_values("FECHA", ascending=True, inplace=True)

        if nombre == "Farmashop" or nombre == "Bela":
            index = ["TIENDA", "DEPARTAMENTO", "Jefe Supervisor"]
        else:
            index = ["TIENDA", "CODIGO DE PAGO", "DEPARTAMENTO", "Jefe Supervisor"]

        df_anexo_for_tabla.to_excel("df_anexo_for_tabla.xlsx")

        df_group = df_anexo_for_tabla.pivot_table(index=index,
                                        columns="FECHA",
                                        values="N° FUNC.",
                                        aggfunc=len)
        df_group.to_excel("df_group.xlsx")

        if nombre == "Farmashop":
            color_tab = "5086C1"
            color_header = "5086C1"
            color_title = "FFFFFF"
        elif nombre == "Bela":
            color_tab = "990C58"
            color_header = "990C58"
            color_title = "FFFFFF"
        elif nombre == "SEO":
            color_tab = "8A6B10"
            color_header = "8A6B10"
            color_title = "FFFFFF"
        elif nombre == "Ecommerce 900":
            color_tab = "204E60"
            color_header = "204E60"
            color_title = "FFFFFF"
        elif nombre == "Logistica":
            color_tab = "992900"
            color_header = "992900"
            color_title = "FFFFFF"

        #---------------------------TABLA---------------------------------------------------
        sheet_name = f'{nombre} ausencias'

        workbook.create_sheet(sheet_name)
        workbook_sheet = workbook[sheet_name]
        workbook_sheet.sheet_properties.tabColor = color_tab

        medidas = df_group.shape

        if nombre == "Farmashop" or nombre == "Bela":
            largo_num = medidas[0] + 2
            ancho_num = medidas[1] + 3
            marcador = True
        else:
            largo_num = medidas[0] + 2
            ancho_num = medidas[1] + 4
            marcador = False

        ancho_letra = get_column_letter(ancho_num)

        for row in dataframe_to_rows(df_group, index=True, header=True):
            workbook_sheet.append(row)

        borde_param = Side(style= 'thin', color= "000000")
        selecciona_celdas(rango= f"A1:{ancho_letra}{largo_num}",
                          sheet= workbook_sheet,
                          font= Font(name= 'Calibri Light'),
                          alignment=Alignment(horizontal= 'center', vertical= 'center'),
                          border= Border(left= borde_param, right= borde_param,
                                                      top= borde_param, bottom= borde_param))

        rule = ColorScaleRule(start_type= "min", start_color= "FFC247",
                              end_type= "max", end_color= "EC563D")
        if marcador == True:
            selecciona_celdas(rango= f"A1:C{largo_num}",
                              sheet= workbook_sheet,
                              font= Font(bold=True))
            workbook_sheet.conditional_formatting.add(f'D3:{ancho_letra}{largo_num}', rule)

        else:
            selecciona_celdas(rango= f"A1:D{largo_num}",
                              sheet= workbook_sheet,
                              font= Font(bold= True))
            workbook_sheet.conditional_formatting.add(f'E3:{ancho_letra}{largo_num}', rule)

        #Formateo de Headeer
        selecciona_celdas(rango= f"A1:{ancho_letra}2",
                          sheet= workbook_sheet,
                          fill= PatternFill(start_color= color_header, fill_type= "solid"),
                          font= Font(bold= True, color= color_title))

        for column in workbook_sheet.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 3)
            nro_columna = column_index_from_string(get_column_letter(column[0].column))
            if nombre == "Farmashop" or nombre == "Bela":
                if nro_columna <= 3:
                    workbook_sheet.column_dimensions[cell.column_letter].width = adjusted_width
                else:
                    workbook_sheet.column_dimensions[cell.column_letter].width = 10
            else:
                if nro_columna <= 4:
                    workbook_sheet.column_dimensions[cell.column_letter].width = adjusted_width
                else:
                    workbook_sheet.column_dimensions[cell.column_letter].width = 10
    #---------------------------------------------------------------------------------------
        # ANEXO BASE

        sheet_name = f"{nombre} Cod. Pago"
        workbook.create_sheet(sheet_name)
        workbook_sheet = workbook[sheet_name]
        workbook_sheet.sheet_properties.tabColor = color_tab

        medidas = df_anexo.shape

        largo_num = medidas[0] + 1
        ancho_num = medidas[1] + 1
        ancho_letra = get_column_letter(ancho_num - 1)

        #Entrada de datos a hoja
        for row in dataframe_to_rows(df_anexo, index= False, header= True):
            workbook_sheet.append(row)

        #Formateo de hoja general
        borde_param = Side(style= 'thin', color= "808080")
        selecciona_celdas(rango= f"A1:{ancho_letra}{largo_num}",
                          sheet= workbook_sheet,
                          font= Font(name='Calibri Light'),
                          alignment= Alignment(horizontal='center', vertical='center'),
                          border= Border(left=borde_param, right=borde_param,
                                                      top=borde_param, bottom=borde_param))
        #Formateo del header
        selecciona_celdas(rango= f"A1:{ancho_letra}1",
                          sheet= workbook_sheet,
                          font= Font(bold= True, color= color_title),
                          fill= PatternFill(start_color=color_header, fill_type="solid"))

        #Ajuste de tamaño de columnas segun celda mas larga
        for column in workbook_sheet.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 3)
            workbook_sheet.column_dimensions[cell.column_letter].width = adjusted_width



    # ------------------------------------ GRAFICAS ----------------------------------------
        if nombre == "Farmashop":
        #--FIG_1-----Tipo de ausencia - Nivel Compañia
            df_fig_1 = df_anexo["Jefe Supervisor"].value_counts().reset_index().copy()
            fig_1 = px.pie(df_fig_1, values='Jefe Supervisor', names='index')
            fig_1.write_image(fr"C:\Users\bchavat\Desktop\Automat Reportes\fig_1_{nombre}.jpg")
        #--FIG_2----- Tipo de ausencia - Nivel Gerencias
            gerentes = df_anexo["Jefe Supervisor"].unique().tolist() # Extrae nombres de gerentes
            df_fig_2 = df_anexo[["Jefe Supervisor", "TIPO"]].copy() # Gerentes y Tipo de ausencia
            for i in gerentes:
                mask_2_1 = df_fig_2.loc[:,"Jefe Supervisor"] == i
                df_fig_2_1 = df_fig_2[mask_2_1].value_counts().reset_index() # Gerente y conteo de cuantas planificadas y no planificadas
                df_fig_2_1 = df_fig_2_1.rename(columns={0: "Valor"})
                fig_2 = px.pie(df_fig_2_1, values='Valor', names='TIPO',
                title=f"Ausencias en gerencia de {i}")
                fig_2.write_image(
                    fr"C:\Users\bchavat\Desktop\Automat Reportes\fig_2_1_{nombre}_{i}.jpg")
        #--FIG_3----- Tipo de Ausencias - Nivel Supervisores
            df_fig_3 = df_anexo[["SUPERVISOR", "TIPO"]].groupby(
                                        ["SUPERVISOR", "TIPO"]).value_counts().reset_index()
            fig_3 = px.bar(df_fig_3, x="SUPERVISOR", y= 0, color="TIPO",
                                        color_discrete_sequence=px.colors.qualitative.Vivid)
            fig_3.write_image(
                fr"C:\Users\bchavat\Desktop\Automat Reportes\fig_3_{nombre}_Supervisores.jpg")
        #--FIG_4----- Tipo de Ausencias - Nivel Sucursales.
            df_fig_4 = df_anexo[["TIENDA", "TIPO"]].value_counts().reset_index()
            df_fig_4 = df_fig_4.rename(columns={0: "Valor"})
            df_fig_4_1 = df_fig_4.groupby(["TIENDA"]).sum().reset_index()
            df_fig_4_1 = pd.merge(left=df_fig_4, right=df_fig_4_1, how="left", on="TIENDA")
            df_fig_4_1 = df_fig_4_1.rename(columns={"Valor_x": "Valor", "Valor_y": "Valor Total"})
            mask_4_1 = df_fig_4_1.loc[:, "Valor Total"] > 9
            df_fig_4_1 = df_fig_4_1[mask_4_1]

            fig_4 = px.bar(df_fig_4_1, x="TIENDA", y="Valor", color="TIPO",
                                        color_discrete_sequence=px.colors.qualitative.Vivid)
            fig_4.write_image(
                fr"C:\Users\bchavat\Desktop\Automat Reportes\fig_4_{nombre}_Sucursales.jpg")
        #--FIG_5-----
            df_fig_5 = df_anexo[df_anexo["CODIGO DE PAGO"].str.contains(r"LIC\w+", regex=True)]
            df_fig_5 = df_fig_5["Jefe Supervisor"].value_counts().reset_index()
            fig_5 = px.pie(df_fig_5, values='Jefe Supervisor', names='index')
            fig_5.update_traces(text=df_fig_5['Jefe Supervisor'], textinfo='text')
            fig_5.write_image(fr"C:\Users\bchavat\Desktop\Automat Reportes\fig_5_{nombre}.jpg")

            print(df_fig_5)

    workbook.remove(workbook['Sheet'])

    workbook.save("anexo_ausencias.xlsx")


# procesa_reporte_ausentismo()









